from __future__ import annotations

import csv
import io
import json
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

from ..util import atomic_write_bytes, sha256_bytes
from .query import query_release

__all__ = ["query_release", "render_analytical_workbook_bundle", "write_analytical_workbook_bundle"]

_CANONICAL_DOCUMENT_TIME = datetime(2000, 1, 1, 0, 0, 0)
_CANONICAL_ZIP_DATE = (1980, 1, 1, 0, 0, 0)
_CANONICAL_COMPRESSION = zipfile.ZIP_STORED
_CANONICAL_EXTERNAL_ATTR = 0o600 << 16


def _sheet_to_csv(rows: list[dict[str, Any]]) -> str:
    if not rows:
        return "column\n"
    buffer = io.StringIO()
    fieldnames = sorted({key for row in rows for key in row})
    writer = csv.DictWriter(buffer, fieldnames=fieldnames, lineterminator="\n")
    writer.writeheader()
    for row in rows:
        writer.writerow({key: row.get(key, "") for key in fieldnames})
    return buffer.getvalue()


def _canonical_zip_info(filename: str) -> zipfile.ZipInfo:
    info = zipfile.ZipInfo(filename=filename, date_time=_CANONICAL_ZIP_DATE)
    info.compress_type = _CANONICAL_COMPRESSION
    info.create_system = 3
    info.create_version = 20
    info.extract_version = 20
    info.flag_bits = 0
    info.internal_attr = 0
    info.external_attr = _CANONICAL_EXTERNAL_ATTR
    info.extra = b""
    info.comment = b""
    return info


def _render_deterministic_zip(entries: list[tuple[str, bytes]]) -> bytes:
    filenames = [filename for filename, _ in entries]
    if len(filenames) != len(set(filenames)):
        raise ValueError("deterministic archive entries must have unique filenames")

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=_CANONICAL_COMPRESSION, allowZip64=True) as archive:
        archive.comment = b""
        for filename, payload in sorted(entries, key=lambda item: item[0]):
            archive.writestr(_canonical_zip_info(filename), payload, compress_type=_CANONICAL_COMPRESSION)
    return buffer.getvalue()


def _canonicalize_zip_payload(payload: bytes) -> bytes:
    with zipfile.ZipFile(io.BytesIO(payload), "r") as source:
        entries = [(info.filename, source.read(info)) for info in source.infolist()]
    return _render_deterministic_zip(entries)


def _render_native_xlsx(query: dict[str, Any]) -> bytes | None:
    try:
        from openpyxl import Workbook  # type: ignore[import-untyped]
        from openpyxl.writer.excel import ExcelWriter  # type: ignore[import-untyped]
    except ImportError:
        return None

    workbook = Workbook()
    workbook.properties.created = _CANONICAL_DOCUMENT_TIME
    workbook.properties.modified = _CANONICAL_DOCUMENT_TIME

    default = workbook.active
    first = True
    for sheet_name in sorted(query["rows"]):
        rows = query["rows"][sheet_name]
        if first:
            worksheet = default
            worksheet.title = sheet_name[:31]
            first = False
        else:
            worksheet = workbook.create_sheet(title=sheet_name[:31])
        if not rows:
            worksheet.append(["column"])
            continue
        fieldnames = sorted({key for row in rows for key in row})
        worksheet.append(fieldnames)
        for row in rows:
            worksheet.append([row.get(key, "") for key in fieldnames])

    if "verification" in workbook.sheetnames:
        verification = workbook["verification"]
    else:
        verification = workbook.create_sheet(title="verification")
        verification.append(["field", "value"])
    verification.append(["format", "openpyxl-native-xlsx"])

    raw = io.BytesIO()
    archive = zipfile.ZipFile(raw, "w", compression=_CANONICAL_COMPRESSION, allowZip64=True)
    try:
        ExcelWriter(workbook, archive).save()
    finally:
        archive.close()
    return _canonicalize_zip_payload(raw.getvalue())


def _render_fallback_bundle(query: dict[str, Any]) -> bytes:
    readme = (
        "NeuroAI analytical workbook bundle (CSV-in-ZIP fallback).\n"
        "Install optional extra 'products' (openpyxl) for native .xlsx.\n"
        f"release_sha256={query['release_sha256']}\n"
    )
    manifest = (
        json.dumps(
            {
                "format": "csv-in-zip-xlsx-fallback",
                "release_sha256": query["release_sha256"],
                "sheets": sorted(query["rows"]),
                "withheld_claims": query["withheld_claims"],
            },
            ensure_ascii=False,
            indent=2,
            sort_keys=True,
        )
        + "\n"
    )
    entries = [
        ("README.txt", readme.encode("utf-8")),
        ("workbook.manifest.json", manifest.encode("utf-8")),
    ]
    entries.extend(
        (f"sheets/{sheet_name}.csv", _sheet_to_csv(query["rows"][sheet_name]).encode("utf-8"))
        for sheet_name in sorted(query["rows"])
    )
    return _render_deterministic_zip(entries)


def render_analytical_workbook_bundle(query: dict[str, Any]) -> bytes:
    """Render a byte-deterministic native XLSX or CSV-in-ZIP fallback."""
    native = _render_native_xlsx(query)
    if native is not None:
        return native
    return _render_fallback_bundle(query)


def _detect_workbook_format(payload: bytes) -> str:
    try:
        with zipfile.ZipFile(io.BytesIO(payload), "r") as archive:
            names = set(archive.namelist())
    except zipfile.BadZipFile as exc:
        raise ValueError("unrecognized analytical workbook package") from exc
    if {"[Content_Types].xml", "xl/workbook.xml"} <= names:
        return "openpyxl-native-xlsx"
    if {"README.txt", "workbook.manifest.json"} <= names:
        return "csv-in-zip-xlsx-fallback"
    raise ValueError("unrecognized analytical workbook package")


def write_analytical_workbook_bundle(query: dict[str, Any], output: Path) -> dict[str, Any]:
    payload = render_analytical_workbook_bundle(query)
    atomic_write_bytes(output, payload)
    return {
        "output": str(output),
        "sha256": sha256_bytes(payload),
        "bytes": len(payload),
        "format": _detect_workbook_format(payload),
        "boundary": "Workbook sheets are deterministic projections of canonical JSON; no manual master data.",
    }
