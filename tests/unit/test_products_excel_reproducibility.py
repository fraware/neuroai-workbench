from __future__ import annotations

import io
import sys
import warnings
import zipfile
from datetime import datetime
from pathlib import Path

import pytest

from neuroai_workbench.products import excel as excel_module

_CANONICAL_ZIP_DATE = (1980, 1, 1, 0, 0, 0)
_CANONICAL_EXTERNAL_ATTR = 0o600 << 16


def _synthetic_query() -> dict[str, object]:
    release_sha256 = "1" * 64
    return {
        "release_sha256": release_sha256,
        "withheld_claims": ["synthetic fixture"],
        "rows": {
            "empty": [],
            "organizations": [
                {
                    "canonical_name": "Example Neurotech",
                    "organization_id": "org-example",
                    "verification_state": "verified",
                }
            ],
            "verification": [
                {"field": "release_sha256", "value": release_sha256},
            ],
        },
    }


def _source_zip(
    entries: list[tuple[str, bytes]],
    *,
    date_time: tuple[int, int, int, int, int, int],
    archive_comment: bytes,
    member_comment: bytes,
    external_attr: int,
) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.comment = archive_comment
        for filename, payload in entries:
            info = zipfile.ZipInfo(filename=filename, date_time=date_time)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 0
            info.external_attr = external_attr
            info.extra = b"\xfe\xca\x00\x00"
            info.comment = member_comment
            archive.writestr(info, payload)
    return buffer.getvalue()


def _assert_canonical_zip_metadata(
    payload: bytes,
    *,
    expected_names: list[str] | None = None,
) -> None:
    with zipfile.ZipFile(io.BytesIO(payload), "r") as archive:
        infos = archive.infolist()
        names = [info.filename for info in infos]
        assert archive.comment == b""
        if expected_names is None:
            assert names == sorted(names)
        else:
            assert names == expected_names
        assert infos
        for info in infos:
            assert info.date_time == _CANONICAL_ZIP_DATE
            assert info.compress_type == zipfile.ZIP_STORED
            assert info.create_system == 3
            assert info.create_version == 20
            assert info.extract_version == 20
            assert info.external_attr == _CANONICAL_EXTERNAL_ATTR
            assert info.extra == b""
            assert info.comment == b""


def test_zip_canonicalization_erases_source_metadata_and_order() -> None:
    entries = [("b.txt", b"bravo"), ("a.txt", b"alpha")]
    first = _source_zip(
        entries,
        date_time=(2020, 1, 2, 3, 4, 6),
        archive_comment=b"first archive",
        member_comment=b"first member",
        external_attr=0o644 << 16,
    )
    second = _source_zip(
        list(reversed(entries)),
        date_time=(2026, 8, 15, 12, 34, 56),
        archive_comment=b"second archive",
        member_comment=b"second member",
        external_attr=0o755 << 16,
    )

    canonical_first = excel_module._canonicalize_zip_payload(first)
    canonical_second = excel_module._canonicalize_zip_payload(second)

    assert first != second
    assert canonical_first == canonical_second
    _assert_canonical_zip_metadata(canonical_first)


def test_zip_canonicalization_rejects_duplicate_members() -> None:
    buffer = io.BytesIO()
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        with zipfile.ZipFile(buffer, "w") as archive:
            archive.writestr("duplicate.txt", b"first")
            archive.writestr("duplicate.txt", b"second")

    with pytest.raises(ValueError, match="unique filenames"):
        excel_module._canonicalize_zip_payload(buffer.getvalue())


def test_native_workbook_core_properties_and_package_metadata_are_canonical() -> None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        pytest.skip("native workbook dependency is not installed")

    query = _synthetic_query()
    payload = excel_module.render_analytical_workbook_bundle(query)
    workbook = load_workbook(io.BytesIO(payload))

    assert workbook.properties.created == datetime(2000, 1, 1)
    assert workbook.properties.modified == datetime(2000, 1, 1)
    assert "verification" in workbook.sheetnames
    assert "verification1" not in workbook.sheetnames
    flat = [item for row in workbook["verification"].iter_rows(values_only=True) for item in row]
    assert query["release_sha256"] in flat
    _assert_canonical_zip_metadata(payload)


def test_fallback_bundle_uses_same_deterministic_archive_contract(monkeypatch: pytest.MonkeyPatch) -> None:
    query = _synthetic_query()
    with monkeypatch.context() as patch:
        patch.setitem(sys.modules, "openpyxl", None)
        first = excel_module.render_analytical_workbook_bundle(query)
        second = excel_module.render_analytical_workbook_bundle(query)

    assert first == second
    assert excel_module._detect_workbook_format(first) == "csv-in-zip-xlsx-fallback"
    _assert_canonical_zip_metadata(
        first,
        expected_names=[
            "README.txt",
            "workbook.manifest.json",
            "sheets/empty.csv",
            "sheets/organizations.csv",
            "sheets/verification.csv",
        ],
    )
    with zipfile.ZipFile(io.BytesIO(first), "r") as archive:
        assert archive.read("sheets/empty.csv") == b"column\n"
        assert query["release_sha256"].encode("utf-8") in archive.read("README.txt")


def test_format_detection_fails_closed_for_unknown_packages() -> None:
    with pytest.raises(ValueError, match="unrecognized analytical workbook package"):
        excel_module._detect_workbook_format(b"not-a-zip")

    unknown = excel_module._render_deterministic_zip([("other.txt", b"payload")])
    with pytest.raises(ValueError, match="unrecognized analytical workbook package"):
        excel_module._detect_workbook_format(unknown)


def test_write_workbook_renders_exactly_once(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    query = _synthetic_query()
    payload = excel_module.render_analytical_workbook_bundle(query)
    calls: list[dict[str, object]] = []

    def render_once(candidate: dict[str, object]) -> bytes:
        calls.append(candidate)
        return payload

    monkeypatch.setattr(excel_module, "render_analytical_workbook_bundle", render_once)
    output = tmp_path / "deterministic.xlsx"
    metadata = excel_module.write_analytical_workbook_bundle(query, output)

    assert calls == [query]
    assert output.read_bytes() == payload
    assert metadata["bytes"] == len(payload)
    assert metadata["format"] == "openpyxl-native-xlsx"
