from __future__ import annotations

import json
from pathlib import Path

from neuroai_workbench.products.excel import render_analytical_workbook_bundle, write_analytical_workbook_bundle
from neuroai_workbench.products.generate import generate_publication_set
from neuroai_workbench.products.query import query_release

REPO = Path(__file__).resolve().parents[2]
COMPACT = REPO / "examples" / "observatory" / "canonical_successor_snapshot_v1.7.json"
FULL = REPO / "examples" / "observatory" / "evidence_depth_release_v1.4.json"


def test_workbook_render_is_deterministic() -> None:
    query = query_release(COMPACT)
    first = render_analytical_workbook_bundle(query)
    second = render_analytical_workbook_bundle(query)
    assert first == second
    assert len(first) > 0


def test_workbook_includes_release_identity() -> None:
    query = query_release(COMPACT)
    payload = render_analytical_workbook_bundle(query)
    try:
        from openpyxl import load_workbook
    except ImportError:
        import io
        import zipfile

        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            assert "sheets/verification.csv" in archive.namelist()
            verification = archive.read("sheets/verification.csv").decode("utf-8")
            assert query["release_sha256"] in verification
        return
    workbook = load_workbook(__import__("io").BytesIO(payload))
    assert "verification" in workbook.sheetnames
    assert "verification1" not in workbook.sheetnames
    flat = [item for row in workbook["verification"].iter_rows(values_only=True) for item in row]
    assert query["release_sha256"] in flat


def test_full_depth_projection_has_richer_columns() -> None:
    summary = query_release(FULL, depth="summary", limit=None)
    full = query_release(FULL, depth="full", limit=None)
    assert set(summary["rows"]["organizations"][0]) == {
        "organization_id",
        "canonical_name",
        "verification_state",
    }
    assert {
        "roles",
        "aliases",
        "countries",
        "regions",
        "evidence_state",
        "claim_boundary",
        "headquarters_country",
        "unesco_region",
    } <= set(full["rows"]["organizations"][0])
    assert {"url", "evidence_state", "claim_boundary", "title"} <= set(full["rows"]["sources"][0])
    assert "aliases" in full["rows"]
    assert {"organization_id", "alias"} <= set(full["rows"]["aliases"][0])
    for sheet in (
        "organization_resolution",
        "regional_expansion",
        "ownership_capital_events",
        "models",
        "models_datasets",
        "trial_sites",
        "participant_authority",
        "suppliers",
        "data_quality_findings",
    ):
        assert full["rows"][sheet], f"expected non-empty sheet {sheet}"
    # Missing canonical sections stay omitted.
    for absent in ("systems", "captures", "candidates", "adjudications", "evidence_register"):
        assert absent not in full["rows"]


def test_full_list_projection_falls_through_empty_canonical_key(tmp_path: Path) -> None:
    """Empty first alias must not block a later non-empty key (Bugbot: silent sheet drop)."""
    release = json.loads(FULL.read_text(encoding="utf-8"))
    models = release.get("representative_model_records")
    assert isinstance(models, list) and models
    release["representative_model_records"] = []
    release["models"] = models
    path = tmp_path / "empty-primary-models-alias.json"
    path.write_text(json.dumps(release), encoding="utf-8")

    query = query_release(path, depth="full", limit=None)
    assert "models" in query["rows"]
    assert query["rows"]["models"]
    assert len(query["rows"]["models"]) == len(models)


def test_full_depth_compact_includes_delta_and_reopening_detail() -> None:
    full = query_release(COMPACT, depth="full", limit=None)
    assert full["rows"]["reopening_decisions"]
    assert "basis" in full["rows"]["reopening_decisions"][0]
    assert "required_actions" in full["rows"]["reopening_decisions"][0]
    assert full["rows"]["delta_records"]
    assert "delta_section" in full["rows"]["delta_records"][0]
    assert full["rows"]["baseline_counts"]
    assert full["rows"]["provenance_links"]


def test_query_preview_limit_and_release_unbounded() -> None:
    preview = query_release(FULL, limit=50)
    release = json.loads(FULL.read_text(encoding="utf-8"))
    assert len(preview["rows"]["organizations"]) == min(50, len(release.get("organizations", [])))
    assert len(preview["rows"]["sources"]) == min(50, len(release.get("sources", [])))
    full = query_release(FULL, limit=None)
    assert len(full["rows"]["organizations"]) == len(release.get("organizations", []))
    assert len(full["rows"]["sources"]) == len(release.get("sources", []))


def test_generate_publication_set_writes_workbook(tmp_path: Path) -> None:
    report = generate_publication_set(COMPACT, tmp_path, limit=None, depth="full")
    workbook_path = tmp_path / "analytical-workbook.xlsx"
    assert workbook_path.is_file()
    assert report["products"]["analytical_workbook"]["sha256"]
    assert report["release_sha256"] == query_release(COMPACT, depth="full", limit=None)["release_sha256"]
    assert (tmp_path / "current-state-report.docx").is_file()
    assert (tmp_path / "current-state-report.pdf").is_file()
    try:
        from openpyxl import load_workbook
    except ImportError:
        return
    workbook = load_workbook(workbook_path)
    assert "verification" in workbook.sheetnames
    assert "verification1" not in workbook.sheetnames


def test_generate_publication_set_rejects_summary_or_limited() -> None:
    import pytest

    with pytest.raises(ValueError, match="depth='full'"):
        generate_publication_set(COMPACT, Path("."), depth="summary", limit=None)
    with pytest.raises(ValueError, match="limit=None"):
        generate_publication_set(COMPACT, Path("."), depth="full", limit=50)


def test_write_workbook_matches_render(tmp_path: Path) -> None:
    query = query_release(COMPACT)
    output = tmp_path / "book.xlsx"
    meta = write_analytical_workbook_bundle(query, output)
    assert output.is_file()
    assert meta["bytes"] == output.stat().st_size
    assert meta["format"] in {"openpyxl-native-xlsx", "csv-in-zip-xlsx-fallback"}
    if meta["format"] == "csv-in-zip-xlsx-fallback":
        import zipfile

        with zipfile.ZipFile(output) as archive:
            assert "sheets/verification.csv" in archive.namelist()
        return
    from openpyxl import load_workbook

    workbook = load_workbook(output)
    assert "verification" in workbook.sheetnames
